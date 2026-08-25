"""CSV / TSV / XLSX extraction.

Tabular files are rendered as SHAPE, not contents: columns, the true row count,
and a bounded sample. A marketing export can be tens of thousands of rows; an
agent asked "what's in this?" needs the columns and a feel for the values, and
pushing the whole grid at it buries the answer in its own context.

The true row count is always reported even when rows are elided, so an agent can
never mistake a 50-row sample for a 50-row file -- the failure mode that would
have it confidently reason about totals that are wrong.
"""

from __future__ import annotations

import csv
import io

from artemis.files.extract.base import (
    MAX_TABULAR_SAMPLE_ROWS,
    ExtractedFile,
    FileParseError,
    TabularShape,
    cap_text,
)
from artemis.files.extract.text import decode_bytes

# csv.Sniffer needs a bounded sample; the whole file is wasteful and it only ever
# inspects the head anyway.
_SNIFF_BYTES = 16384


def _sniff_delimiter(sample: str, filename: str) -> str:
    """Detect the column separator, preferring the extension when it is explicit.

    The extension wins over sniffing because it is deterministic and the sniffer
    is not. Measured 2026-08-25: `csv.Sniffer` handles delimiters-inside-values
    better than expected (a .tsv with comma-heavy free-text fields still sniffs
    as tab), but it raises outright on a SINGLE-COLUMN file -- there is no
    delimiter to find -- and it only ever inspects the head, so a file whose
    first rows are unrepresentative can still be misread. Trusting the name the
    author chose avoids betting on any of that.

    A plain comma is the last resort rather than an exception: a single-column
    file is perfectly readable, it simply has one column.
    """
    lowered = filename.lower()
    if lowered.endswith((".tsv", ".tab")):
        return "\t"
    if lowered.endswith(".csv"):
        return ","
    try:
        return str(csv.Sniffer().sniff(sample[:_SNIFF_BYTES], delimiters=",\t;|").delimiter)
    except csv.Error:
        return ","


def extract_delimited(
    payload: bytes, *, filename: str, mimetype: str = "text/csv"
) -> ExtractedFile:
    """Extract a CSV/TSV-family file."""
    decoded, encoding_note = decode_bytes(payload, filename=filename)
    delimiter = _sniff_delimiter(decoded, filename)

    try:
        rows = list(csv.reader(io.StringIO(decoded), delimiter=delimiter))
    except csv.Error as exc:
        raise FileParseError(
            f"{filename} looks like a delimited table but could not be parsed ({exc})."
        ) from exc

    rows = [r for r in rows if any(cell.strip() for cell in r)]
    if not rows:
        raise FileParseError(f"{filename} parsed as a table but contains no rows.")

    header, *body = rows
    columns = [c.strip() for c in header]
    sample = body[:MAX_TABULAR_SAMPLE_ROWS]

    shape = TabularShape(
        columns=columns,
        total_rows=len(body),
        sample_rows=sample,
        truncated=len(body) > len(sample),
    )
    text, truncated = cap_text(_render_table(shape, filename))

    notes = [n for n in (encoding_note,) if n]
    if delimiter != ",":
        notes.append(f"Parsed with {'tab' if delimiter == chr(9) else repr(delimiter)} separator.")

    return ExtractedFile(
        filename=filename,
        mimetype=mimetype,
        kind="tabular",
        text=text,
        size_bytes=len(payload),
        tables=[shape],
        truncated=truncated or shape.truncated,
        notes=notes,
    )


def extract_xlsx(payload: bytes, *, filename: str, mimetype: str = "") -> ExtractedFile:
    """Extract an Excel workbook, every sheet.

    Reads with `data_only=True` so formula cells yield their last computed VALUE
    rather than the formula string -- an agent asked about a total wants 48213,
    not "=SUM(B2:B900)". The tradeoff is real and worth stating: a workbook saved
    by a tool that never computed the formulas has no cached values, and those
    cells read as empty. That is flagged in the notes rather than passed off as
    genuinely blank data.
    """
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            io.BytesIO(payload), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:  # openpyxl raises a wide variety on malformed input
        raise FileParseError(
            f"{filename} could not be opened as a spreadsheet ({type(exc).__name__}). "
            "If it is password-protected, it must be unlocked before it can be read."
        ) from exc

    shapes: list[TabularShape] = []
    empty_formula_cells = False
    try:
        for sheet in workbook.worksheets:
            rows_iter = sheet.iter_rows(values_only=True)
            raw_rows = [
                ["" if cell is None else str(cell) for cell in row]
                for row in rows_iter
                if row is not None and any(cell is not None for cell in row)
            ]
            if not raw_rows:
                continue
            header, *body = raw_rows
            sample = body[:MAX_TABULAR_SAMPLE_ROWS]
            if any(not cell for row in sample for cell in row):
                empty_formula_cells = True
            shapes.append(
                TabularShape(
                    columns=[c.strip() for c in header],
                    total_rows=len(body),
                    sample_rows=sample,
                    sheet_name=str(sheet.title),
                    truncated=len(body) > len(sample),
                )
            )
    finally:
        workbook.close()

    if not shapes:
        raise FileParseError(f"{filename} opened as a spreadsheet but every sheet is empty.")

    rendered = "\n\n".join(_render_table(shape, filename) for shape in shapes)
    text, truncated = cap_text(rendered)

    notes: list[str] = []
    if len(shapes) > 1:
        notes.append(
            f"Workbook has {len(shapes)} sheets: {', '.join(s.sheet_name or '?' for s in shapes)}."
        )
    if empty_formula_cells:
        notes.append(
            "Some cells are blank. Formula cells show their last SAVED value, so a "
            "workbook whose formulas were never computed can read as empty here."
        )

    return ExtractedFile(
        filename=filename,
        mimetype=mimetype or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        kind="tabular",
        text=text,
        size_bytes=len(payload),
        tables=shapes,
        truncated=truncated or any(s.truncated for s in shapes),
        notes=notes,
    )


def _render_table(shape: TabularShape, filename: str) -> str:
    """Render a shape as text an agent reads.

    States the real row count up front and, when rows were elided, says so
    explicitly in the body -- not only in a metadata field the model may not be
    shown.
    """
    lines: list[str] = []
    title = f"{filename}" + (f" - sheet '{shape.sheet_name}'" if shape.sheet_name else "")
    lines.append(f"{title}: {shape.total_rows:,} data rows, {len(shape.columns)} columns")
    lines.append("Columns: " + " | ".join(shape.columns))
    if shape.sample_rows:
        shown = len(shape.sample_rows)
        label = (
            f"First {shown} of {shape.total_rows:,} rows"
            if shape.truncated
            else f"All {shown} rows"
        )
        lines.append(f"{label}:")
        lines.append(" | ".join(shape.columns))
        for row in shape.sample_rows:
            lines.append(" | ".join(row))
    if shape.truncated:
        lines.append(
            f"[{shape.total_rows - len(shape.sample_rows):,} further rows not shown -- "
            "totals or counts must not be computed from this sample.]"
        )
    return "\n".join(lines)
